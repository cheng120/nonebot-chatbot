#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path


def _load_json(path: Path) -> dict:
	with path.open("r", encoding="utf-8") as f:
		return json.load(f)


def _save_json(path: Path, data: dict) -> None:
	tmp = path.with_suffix(path.suffix + ".tmp")
	with tmp.open("w", encoding="utf-8") as f:
		json.dump(data, f, ensure_ascii=False, indent=2, sort_keys=True)
		f.write("\n")
	tmp.replace(path)


def _norm(s: str) -> str:
	return "".join((s or "").strip().split()).lower()


def _pick_col(headers: list[str], candidates: list[str]) -> str | None:
	hmap = {_norm(h): h for h in headers}
	for c in candidates:
		if _norm(c) in hmap:
			return hmap[_norm(c)]
	return None


def _read_xlsx_rows(xlsx: Path, sheet: str | None) -> list[dict]:
	try:
		from openpyxl import load_workbook
	except Exception as e:  # pragma: no cover
		raise RuntimeError(
			"缺少依赖 openpyxl。请先安装：pip install openpyxl"
		) from e

	wb = load_workbook(filename=str(xlsx), read_only=True, data_only=True)
	ws = wb[sheet] if sheet else wb.active
	rows = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows)
	except StopIteration:
		return []
	headers = [str(h).strip() if h is not None else "" for h in header_row]

	out: list[dict] = []
	for r in rows:
		item: dict = {}
		for i, v in enumerate(r):
			if i >= len(headers):
				continue
			k = headers[i]
			if not k:
				continue
			item[k] = v
		if item:
			out.append(item)
	return out


def _build_mapping(rows: list[dict]) -> tuple[dict[str, str], dict[str, str]]:
	if not rows:
		return {}, {}
	headers = list(rows[0].keys())
	name_col = _pick_col(headers, ["装备名", "名称", "name", "item", "装备名称"])
	desc_col = _pick_col(headers, ["详细描述", "描述", "desc", "description", "说明", "介绍"])
	id_col = _pick_col(headers, ["ID", "id", "编号", "装备ID"])

	if not desc_col:
		raise RuntimeError("Excel 中找不到“描述/详细描述”列（支持：描述/详细描述/description/desc/说明/介绍）")

	by_name: dict[str, str] = {}
	by_id: dict[str, str] = {}
	for r in rows:
		desc = r.get(desc_col)
		if desc is None:
			continue
		desc_s = str(desc).strip()
		if not desc_s:
			continue
		if name_col and r.get(name_col) is not None:
			by_name[_norm(str(r.get(name_col)))] = desc_s
		if id_col and r.get(id_col) is not None:
			by_id[str(r.get(id_col)).strip()] = desc_s
	return by_name, by_id


def _update_equipment_json(json_path: Path, by_name: dict[str, str], by_id: dict[str, str], field: str) -> tuple[int, int]:
	data = _load_json(json_path)
	if not isinstance(data, dict):
		raise RuntimeError(f"{json_path} 不是对象结构 JSON")

	updated = 0
	total = 0
	for k, v in data.items():
		if not isinstance(v, dict):
			continue
		total += 1
		desc = None
		if k in by_id:
			desc = by_id[k]
		else:
			n = v.get("name")
			if isinstance(n, str) and _norm(n) in by_name:
				desc = by_name[_norm(n)]
		if desc is None:
			continue
		if v.get(field) == desc:
			continue
		v[field] = desc
		updated += 1

	if updated:
		_save_json(json_path, data)
	return updated, total


def main(argv: list[str]) -> int:
	parser = argparse.ArgumentParser(description="将 Excel 中的装备详细描述同步到修仙装备 JSON")
	parser.add_argument("--xlsx", required=True, help="Excel 路径")
	parser.add_argument("--sheet", default=None, help="工作表名称（默认：第一个）")
	parser.add_argument("--field", default="desc", help="写入 JSON 的字段名（默认：desc）")
	parser.add_argument(
		"--json",
		action="append",
		default=[],
		help="要更新的 JSON 文件（可重复）。默认更新 data/xiuxian/装备/法器.json 与 防具.json",
	)
	args = parser.parse_args(argv)

	root = Path(__file__).resolve().parents[1]
	xlsx = Path(args.xlsx).expanduser()
	if not xlsx.is_absolute():
		xlsx = (Path.cwd() / xlsx).resolve()
	if not xlsx.exists():
		raise RuntimeError(f"找不到 Excel：{xlsx}")

	json_files = [Path(p) for p in args.json] if args.json else [
		root / "data" / "xiuxian" / "装备" / "法器.json",
		root / "data" / "xiuxian" / "装备" / "防具.json",
	]

	rows = _read_xlsx_rows(xlsx, args.sheet)
	by_name, by_id = _build_mapping(rows)
	if not by_name and not by_id:
		raise RuntimeError("Excel 未解析到可用数据（需要至少：名称/装备名 + 描述/详细描述；或 ID + 描述/详细描述）")

	for jf in json_files:
		jp = jf if jf.is_absolute() else (root / jf)
		u, t = _update_equipment_json(jp, by_name, by_id, args.field)
		print(f"{jp}: 更新 {u}/{t}")
	return 0


if __name__ == "__main__":
	try:
		raise SystemExit(main(sys.argv[1:]))
	except Exception as e:
		print(f"错误：{e}", file=sys.stderr)
		raise SystemExit(2)

